#!/usr/bin/env python3
import openpyxl
from pathlib import Path
from pensPop import pensMember

class pensMort(object):
    def __init__(self):
        m_file = Path('..', 'mortalityTables', 'pub-2010-amount-mort-rates.xlsx')
        m_wb = openpyxl.load_workbook(m_file)
        pubG = m_wb['PubG-2010']
        self.data_f = {}
        self.data_m = {}
        
        
        for i, row in enumerate(pubG.iter_rows(5, pubG.max_row)):
            if i == 0:
                self.data_f['Employee'] = []
                self.data_f['Healthy Retiree'] = []
                self.data_f['Disabled Retiree'] = []
                self.data_f['Contingent Survivor'] = []

                self.data_m['Employee'] = []
                self.data_m['Healthy Retiree'] = []
                self.data_m['Disabled Retiree'] = []
                self.data_m['Contingent Survivor'] = []

            else:
                self.data_f['Employee'].append(row[3])
                self.data_f['Healthy Retiree'].append(row[4])
                self.data_f['Disabled Retiree'].append(row[5])
                self.data_f['Contingent Survivor'].append(row[6])

                self.data_m['Employee'].append(row[8])
                self.data_m['Healthy Retiree'].append(row[9])
                self.data_m['Disabled Retiree'].append(row[10])
                self.data_m['Contingent Survivor'].append(row[11])
				
    def getRate(sex, age, status):
        age -= 18
        if sex == "F":
            table = self.data_f
        else:
            table = self.data_m
		
        if status == "active" or status == "separated":
            rate = table['Employee'][age]
        elif status == "retired":
            rate = table['Healthy Retiree'][age]
        elif status == "deceased":
            rate = 1  
        else:
            rate = 0
			
        return rate

if __name__ == '__main__':
	m1 = pensMember(18, "F", 1, 600, 2010)
	m2 = pensMember(55, "M", 20, 2000, 2010)
	
	r1 = m1.mRates.getRate(m1.sex, m1.age, m1.status)
	r2 = m2.mRates.getRate(m2.sex, m2.age, m2.status)
	
	if r1 == 0.00013 and r2 == 0.00219:
		print("pensMort test complete, everything checks out!")
	else:
		print("Error. r1 = " + str(r1) + " instead of 0.00013, r2 = " + str(r2) + " instead of 0.00219.")